"""
Extracción y limpieza determinística del corpus CODEFEST AD ASTRA 2026.

Lee el inventario maestro (Indice_Datos_Codefest.xlsx -> hoja "Inventario de Archivos"),
resuelve cada fila contra el disco, extrae texto según el tipo de archivo (PDF, JSON, CSV,
Excel, Texto, Imagen, Otro) y escribe un JSONL por fenómeno en data/clean/, separando los
documentos procesados de los pendientes (excluidos, no descargados o con error).

No usa ningún modelo generativo: toda la extracción es determinística (parsers + OCR clásico).
"""

from __future__ import annotations

import argparse
import csv
import gc
import hashlib
import io
import json
import os
import re
import sys
import tempfile
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import blackboxprotobuf
import chardet
import openpyxl
import pdfplumber
import psutil
from langdetect import DetectorFactory, LangDetectException, detect
from pdf2image import convert_from_path, pdfinfo_from_path
from pdf2image.exceptions import PDFInfoNotInstalledError, PDFPageCountError, PopplerNotInstalledError
from PIL import Image
import pytesseract

# langdetect no es determinístico entre corridas si no se fija la semilla
DetectorFactory.seed = 0

# ---------------------------------------------------------------------------
# Configuración general
# ---------------------------------------------------------------------------

RUTA_INDICE = Path("Indice_Datos_Codefest.xlsx")
RAIZ_DATOS = Path(".")  # esta carpeta hace las veces de F3_Dinamicas_Territoriales
CARPETA_SALIDA = Path("data/clean")

# Patrón de exclusión de catálogos/registros de metadata, validado manualmente contra
# la lista completa de nombres de F3 antes de aprobarse (ver conversación de diseño).
# Deliberadamente NO incluye "index" ni "inventario" porque generan falsos positivos
# graves (p.ej. excluirían las publicaciones reales de Stanford AI Index en F1).
PATRON_CATALOGO = re.compile(r"catalog|catalogo|_registro", re.IGNORECASE)

FENOMENOS = ("F1", "F2", "F3")


class OCRNoDisponible(Exception):
    """Se lanza cuando falta el binario de Tesseract o Poppler en el sistema: se distingue
    de un error de extracción real para que el documento quede en pendientes con un motivo
    específico (reprocesable después, sin tocar el resto del pipeline)."""


# Límites duros de OCR para que un PDF grande no pueda saturar memoria: dpi bajo,
# renderizado página por página (nunca el documento completo de una vez) y un tope
# de páginas por documento.
# DPI bajado de 150 a 100 tras los MemoryError reproducidos con CEOBS, RESDAL y
# MAPP_OEA: a 100 dpi el OCR de texto normal sigue siendo legible y el consumo de RAM
# por página baja proporcionalmente al cuadrado del dpi (150->100 es ~2.25x menos RAM).
DPI_OCR = 100
MAX_PAGINAS_OCR = 30

# Cada cuántos documentos procesados se fuerza una recolección de basura.
INTERVALO_GC = 15

# Umbral de RAM del sistema (%) por encima del cual se pausa y se fuerza gc.collect()
# antes de seguir; si tras la pausa se sigue por encima del umbral crítico, el documento
# se salta a pendientes en vez de arriesgarse a tumbar el sistema operativo.
UMBRAL_MEMORIA_PAUSA = 85.0
UMBRAL_MEMORIA_CRITICO = 92.0
SEGUNDOS_PAUSA_MEMORIA = 3


class MemoriaInsuficiente(Exception):
    """Se lanza cuando el uso de RAM del sistema sigue por encima del umbral crítico
    incluso después de pausar y forzar recolección de basura: el documento se salta a
    pendientes con motivo "memoria_insuficiente" en vez de arriesgar un MemoryError o
    tumbar el sistema operativo completo."""


def verificar_memoria_o_pausar() -> None:
    """
    Chequeo manual de memoria vía psutil (equivalente portable a `resource` en Windows,
    donde RLIMIT_AS no está disponible). Se llama antes de procesar cada documento:

    - Si el uso de RAM supera UMBRAL_MEMORIA_PAUSA, pausa unos segundos y fuerza
      gc.collect() para darle tiempo al sistema a liberar memoria de objetos ya
      recolectables antes de seguir.
    - Si después de eso sigue por encima de UMBRAL_MEMORIA_CRITICO, lanza
      MemoriaInsuficiente: el documento se salta a pendientes en vez de intentarlo.
    """
    uso = psutil.virtual_memory().percent
    if uso < UMBRAL_MEMORIA_PAUSA:
        return

    print(f"  [MEMORIA] uso={uso:.1f}% > {UMBRAL_MEMORIA_PAUSA}%: pausando {SEGUNDOS_PAUSA_MEMORIA}s y liberando...", flush=True)
    time.sleep(SEGUNDOS_PAUSA_MEMORIA)
    gc.collect()

    uso_tras_pausa = psutil.virtual_memory().percent
    if uso_tras_pausa >= UMBRAL_MEMORIA_CRITICO:
        raise MemoriaInsuficiente(
            f"uso de RAM sigue en {uso_tras_pausa:.1f}% (>= {UMBRAL_MEMORIA_CRITICO}%) tras pausa y gc.collect()"
        )

# Motivos de "pendiente" que son transitorios (dependen del entorno, no del documento):
# NO se marcan como completados en el checkpoint, para que la próxima corrida los
# reintente automáticamente en vez de darlos por perdidos para siempre.
MOTIVOS_REINTENTABLES = {"ocr_no_disponible", "memoria_insuficiente"}


# ---------------------------------------------------------------------------
# Estructuras de datos
# ---------------------------------------------------------------------------

@dataclass
class FilaInventario:
    """Una fila cruda de la hoja 'Inventario de Archivos'."""
    fenomeno: str
    observatorio: str
    nombre_estandarizado: str
    carpeta: str
    tipo: str


@dataclass
class DocumentoLimpio:
    """Registro final que se escribe en <fenomeno>_documentos.jsonl."""
    doc_id: str
    fuente: str
    formato: str
    fenomeno: str
    texto_limpio: str
    num_caracteres: int
    idioma_detectado: str | None = None
    posible_duplicado: bool | None = None
    hash_contenido: str | None = None

    def to_dict(self) -> dict:
        return {k: v for k, v in self.__dict__.items() if v is not None}


@dataclass
class DocumentoPendiente:
    """Registro para <fenomeno>_pendientes.jsonl: nunca se mezcla con los procesados."""
    doc_id: str
    fuente: str
    formato: str
    fenomeno: str
    motivo: str
    detalle: str = ""

    def to_dict(self) -> dict:
        return self.__dict__


# ---------------------------------------------------------------------------
# Lectura del inventario maestro
# ---------------------------------------------------------------------------

def leer_inventario(ruta_indice: Path = RUTA_INDICE) -> list[FilaInventario]:
    """Lee la hoja 'Inventario de Archivos' y devuelve una fila por documento indexado."""
    wb = openpyxl.load_workbook(ruta_indice, data_only=True)
    ws = wb["Inventario de Archivos"]
    filas = list(ws.iter_rows(values_only=True))
    encabezado = filas[0]
    idx = {nombre: i for i, nombre in enumerate(encabezado)}

    resultado = []
    for fila in filas[1:]:
        if fila[idx["Fenómeno"]] is None:
            continue  # filas vacías al final de la hoja
        resultado.append(
            FilaInventario(
                fenomeno=fila[idx["Fenómeno"]],
                observatorio=fila[idx["Observatorio"]],
                nombre_estandarizado=fila[idx["Nombre estandarizado"]],
                carpeta=fila[idx["Carpeta"]],
                tipo=fila[idx["Tipo"]],
            )
        )
    return resultado


def resolver_ruta_fisica(fila: FilaInventario, raiz: Path = RAIZ_DATOS) -> Path | None:
    """
    Intenta ubicar el archivo real en disco para una fila del inventario.

    Prueba dos variantes porque la columna "Carpeta" incluye como primer segmento el
    nombre de la carpeta del fenómeno (p.ej. "F3_Dinamicas_Territoriales/..."), pero en
    este entorno la raíz de trabajo YA ES esa carpeta para F3 (verificado: los 888
    archivos de F3 resuelven al 100% quitando ese primer segmento). Para F1/F2, cuyas
    carpetas de fenómeno no existen todavía en disco, ambas variantes fallarán y el
    archivo se marca como no descargado.
    """
    candidatos = [
        raiz / fila.carpeta / fila.nombre_estandarizado,
    ]
    partes = fila.carpeta.split("/", 1)
    if len(partes) == 2:
        candidatos.append(raiz / partes[1] / fila.nombre_estandarizado)

    for candidato in candidatos:
        if candidato.is_file():
            return candidato
    return None


def seleccionar_muestra_mixta(filas: list[FilaInventario], n: int) -> list[FilaInventario]:
    """
    Selecciona hasta n filas repartidas proporcionalmente entre los tipos presentes, en
    vez de tomar las primeras n del inventario (que en F3 son todas JSON consecutivos:
    ver ALERTAS_*.json). Así la prueba pequeña sí ejercita PDF/JSON/CSV/Otro(.pbf) juntos.
    """
    por_tipo: dict[str, list[FilaInventario]] = {}
    for fila in filas:
        por_tipo.setdefault(fila.tipo, []).append(fila)

    tipos = list(por_tipo.keys())
    if not tipos:
        return []

    cupo_por_tipo = max(1, n // len(tipos))
    muestra = []
    for tipo in tipos:
        muestra.extend(por_tipo[tipo][:cupo_por_tipo])

    for tipo in tipos:
        if len(muestra) >= n:
            break
        restantes = por_tipo[tipo][cupo_por_tipo:]
        faltan = n - len(muestra)
        muestra.extend(restantes[:faltan])

    return muestra[:n]


def es_catalogo(nombre_estandarizado: str) -> bool:
    """Detecta archivos de catálogo/registro de metadata por patrón de nombre."""
    return bool(PATRON_CATALOGO.search(nombre_estandarizado))


# ---------------------------------------------------------------------------
# Utilidades comunes
# ---------------------------------------------------------------------------

def detectar_idioma(texto: str) -> str | None:
    """Detecta el idioma de un texto; devuelve None si el texto es demasiado corto o ambiguo."""
    texto_util = texto.strip()
    if len(texto_util) < 20:
        return None
    try:
        return detect(texto_util)
    except LangDetectException:
        return None


def calcular_hash(texto: str) -> str:
    """Hash estable del contenido, usado para marcar posibles duplicados (p.ej. tiles repetidos)."""
    return hashlib.sha256(texto.encode("utf-8", errors="ignore")).hexdigest()


def quitar_cabeceras_pies_repetidos(paginas: list[str]) -> list[str]:
    """
    Elimina líneas que se repiten idénticas en la mayoría de páginas de un mismo PDF
    (cabeceras/pies de página institucionales, numeración, etc.), sin depender de un
    formato fijo: se basa en frecuencia de la línea entre páginas.
    """
    if len(paginas) < 3:
        return paginas  # con pocas páginas no hay señal suficiente para detectar patrones

    conteo_lineas: dict[str, int] = {}
    lineas_por_pagina = []
    for texto_pagina in paginas:
        lineas = [l.strip() for l in texto_pagina.split("\n")]
        lineas_por_pagina.append(lineas)
        for linea in set(lineas):
            if linea:
                conteo_lineas[linea] = conteo_lineas.get(linea, 0) + 1

    umbral = max(3, int(len(paginas) * 0.6))
    lineas_repetidas = {linea for linea, veces in conteo_lineas.items() if veces >= umbral}

    paginas_limpias = []
    for lineas in lineas_por_pagina:
        lineas_filtradas = [l for l in lineas if l not in lineas_repetidas]
        paginas_limpias.append("\n".join(lineas_filtradas))
    return paginas_limpias


# ---------------------------------------------------------------------------
# Extractores por tipo de archivo
# ---------------------------------------------------------------------------

def extraer_pdf(ruta: Path) -> str:
    """
    Extrae texto de un PDF página por página, limpia cabeceras/pies repetidos y aplica
    OCR de respaldo (vía render de página a imagen) si el texto nativo viene vacío,
    lo que indica un PDF escaneado sin capa de texto.
    """
    textos_pagina = []
    with pdfplumber.open(ruta) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text() or ""
            textos_pagina.append(texto)

    if all(not t.strip() for t in textos_pagina):
        textos_pagina = _ocr_pdf_escaneado(ruta)

    textos_pagina = quitar_cabeceras_pies_repetidos(textos_pagina)
    return "\n\n".join(t for t in textos_pagina if t.strip())


def _ocr_pdf_escaneado(ruta: Path) -> list[str]:
    """
    Renderiza el PDF a imagen y aplica OCR, página por página, para evitar picos de memoria:

    - dpi=150 (no el default de pdf2image, que es mucho más pesado en RAM por página).
    - Una página a la vez vía first_page/last_page, cerrando la imagen (img.close())
      apenas se le aplica OCR, en vez de convert_from_path completo de una sola vez.
    - Tope de MAX_PAGINAS_OCR páginas por documento; si el PDF tiene más, se trunca y
      se registra en consola para que quede constancia de qué documentos quedaron parciales.
    """
    try:
        info = pdfinfo_from_path(str(ruta))
    except (PDFInfoNotInstalledError, PDFPageCountError, PopplerNotInstalledError) as error:
        raise OCRNoDisponible(f"Poppler no disponible: {error}") from error

    num_paginas = info.get("Pages", 0)
    paginas_a_procesar = min(num_paginas, MAX_PAGINAS_OCR)

    if num_paginas > MAX_PAGINAS_OCR:
        print(
            f"  [OCR] {ruta.name}: truncado a {MAX_PAGINAS_OCR} de {num_paginas} páginas "
            f"(límite MAX_PAGINAS_OCR)", flush=True,
        )

    textos = []
    for num_pagina in range(1, paginas_a_procesar + 1):
        verificar_memoria_o_pausar()
        imagenes = convert_from_path(
            str(ruta), dpi=DPI_OCR, first_page=num_pagina, last_page=num_pagina
        )
        imagen = imagenes[0]
        try:
            textos.append(pytesseract.image_to_string(imagen, lang="spa+eng"))
        except pytesseract.TesseractNotFoundError as error:
            raise OCRNoDisponible(f"Tesseract no disponible: {error}") from error
        finally:
            imagen.close()
            del imagen
            imagenes.clear()
            del imagenes
            gc.collect()
    return textos


def extraer_json(ruta: Path) -> tuple[str, dict]:
    """
    Extrae el texto de un JSON separando cuerpo de metadata de forma dinámica.

    - Si la raíz es una lista, se marca como "bulk dump" en vez de tratarse como un
      documento único (no se intenta fusionar N registros en un solo texto_limpio).
    - Los campos cuyo nombre sugiere metadata (url, autor(es), fecha, tag(s)) se separan
      del cuerpo textual (título, resumen, cuerpo, texto, contenido, descripción), sin
      asumir un esquema fijo por observatorio.

    Devuelve (texto_limpio, metadata_extra), donde metadata_extra incluye
    "es_bulk_dump": True/False.
    """
    with open(ruta, "r", encoding="utf-8", errors="replace") as fh:
        datos = json.load(fh)

    if isinstance(datos, list):
        # Bulk dump: no es un documento individual, se marca y se concatena solo
        # como referencia de tamaño, sin mezclar registros como si fueran un texto.
        return "", {"es_bulk_dump": True, "num_registros": len(datos)}

    if not isinstance(datos, dict):
        return str(datos), {"es_bulk_dump": False}

    claves_metadata = re.compile(r"^(url|autor|autores|fecha|tag|tags|source|link)$", re.IGNORECASE)
    claves_texto = re.compile(r"^(titulo|título|title|resumen|summary|cuerpo|body|texto|text|contenido|content|descripcion|descripción)$", re.IGNORECASE)

    partes_texto = []
    metadata = {}
    for clave, valor in datos.items():
        if isinstance(valor, (dict, list)):
            continue  # estructuras anidadas complejas no se intentan aplanar aquí
        if claves_metadata.match(str(clave)):
            metadata[clave] = valor
        elif claves_texto.match(str(clave)) or isinstance(valor, str):
            if isinstance(valor, str) and valor.strip():
                partes_texto.append(valor)

    metadata["es_bulk_dump"] = False
    return "\n\n".join(partes_texto), metadata


def extraer_csv(ruta: Path) -> str:
    """
    Extrae el contenido textual de un CSV soportando múltiples encodings y separadores.

    Estrategia de fallback robusta (no try/except anidados frágiles):
    1. Detecta el encoding real con chardet sobre una muestra de bytes.
    2. Detecta el separador con csv.Sniffer sobre una muestra de texto ya decodificada.
    3. Si el Sniffer falla, prueba una lista explícita de separadores candidatos en orden
       y se queda con el que produzca más de una columna de forma consistente.
    """
    with open(ruta, "rb") as fh:
        muestra_bytes = fh.read(65536)
    encoding_detectado = chardet.detect(muestra_bytes)["encoding"] or "utf-8"

    with open(ruta, "r", encoding=encoding_detectado, errors="replace") as fh:
        contenido = fh.read()

    separador = _detectar_separador(contenido)

    filas = list(csv.reader(io.StringIO(contenido), delimiter=separador))
    if not filas:
        return ""

    encabezado, registros = filas[0], filas[1:]
    lineas_texto = []
    for registro in registros:
        pares = [f"{col}: {val}" for col, val in zip(encabezado, registro) if val]
        lineas_texto.append(" | ".join(pares))
    return "\n".join(lineas_texto)


def _detectar_separador(contenido: str) -> str:
    """Detecta el separador de un CSV con Sniffer y, si falla, con conteo de candidatos."""
    muestra = contenido[:4096]
    try:
        return csv.Sniffer().sniff(muestra, delimiters=",;\t").delimiter
    except csv.Error:
        pass

    candidatos = [",", ";", "\t"]
    conteos = {sep: muestra.count(sep) for sep in candidatos}
    mejor_separador = max(conteos, key=conteos.get)
    if conteos[mejor_separador] == 0:
        return ","  # sin señal de ningún separador: valor por defecto razonable
    return mejor_separador


def extraer_excel(ruta: Path) -> str:
    """Extrae el contenido textual de todas las hojas de un archivo Excel, celda por celda."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    partes = []
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        for fila in ws.iter_rows(values_only=True):
            valores = [str(v) for v in fila if v is not None]
            if valores:
                partes.append(" | ".join(valores))
    return "\n".join(partes)


def extraer_texto_plano(ruta: Path) -> str:
    """Lee un archivo .txt detectando encoding con chardet."""
    with open(ruta, "rb") as fh:
        contenido_bytes = fh.read()
    encoding_detectado = chardet.detect(contenido_bytes)["encoding"] or "utf-8"
    return contenido_bytes.decode(encoding_detectado, errors="replace")


def extraer_imagen(ruta: Path) -> str:
    """Aplica OCR a una imagen. Si no hay texto legible, devuelve cadena vacía (comportamiento esperado)."""
    with Image.open(ruta) as imagen:
        try:
            return pytesseract.image_to_string(imagen, lang="spa+eng").strip()
        except pytesseract.TesseractNotFoundError as error:
            raise OCRNoDisponible(f"Tesseract no disponible: {error}") from error


def _decodificar_varints_empaquetados(buf: bytes) -> list[int]:
    """Decodifica una secuencia de varints protobuf empaquetados (LEB128 sin signo)."""
    valores = []
    i = 0
    while i < len(buf):
        resultado = 0
        desplazamiento = 0
        while True:
            byte = buf[i]
            i += 1
            resultado |= (byte & 0x7F) << desplazamiento
            if not (byte & 0x80):
                break
            desplazamiento += 7
        valores.append(resultado)
    return valores


def _bytes_o_dict_a_str(campo) -> str:
    """Normaliza un campo protobuf schemaless que blackboxprotobuf puede devolver como
    bytes crudos (caso normal) o, cuando esos bytes parecen a su vez un mensaje protobuf
    válido (ambigüedad inherente a decodificar sin esquema fijo), como dict ya "decodificado"
    un nivel de más. Si ya viene como str, se usa tal cual."""
    if isinstance(campo, bytes):
        return campo.decode("utf-8", errors="replace")
    if isinstance(campo, str):
        return campo
    if isinstance(campo, dict):
        return str(campo)
    return str(campo)


def _valor_mvt_a_texto(valor_crudo: dict) -> str | int | float | bool | None:
    """Traduce un Tile.Value (oneof por número de campo) al tipo Python correspondiente."""
    if "1" in valor_crudo:  # string_value
        return _bytes_o_dict_a_str(valor_crudo["1"])
    if "2" in valor_crudo:  # float_value
        return valor_crudo["2"]
    if "3" in valor_crudo:  # double_value
        return valor_crudo["3"]
    if "4" in valor_crudo:  # int_value
        return valor_crudo["4"]
    if "5" in valor_crudo:  # uint_value
        return valor_crudo["5"]
    if "6" in valor_crudo:  # sint_value
        return valor_crudo["6"]
    if "7" in valor_crudo:  # bool_value
        return bool(valor_crudo["7"])
    return None


def extraer_otro(ruta: Path) -> tuple[str, bool]:
    """
    Maneja archivos clasificados como "Otro" en el inventario. En este corpus corresponden
    a vector tiles Mapbox (.pbf, mayoría en F3) y a imágenes .avif (F2). Los .pbf SÍ se
    decodifican: se extraen las "properties" de cada feature de cada capa y se unen como
    texto "atributo: valor". Se calcula un hash del texto para poder detectar más adelante,
    en la etapa de chunking, contenido repetido entre distintos niveles de zoom del mismo
    área geográfica.

    Se decodifica el protobuf a bajo nivel con `blackboxprotobuf` (sin esquema fijo) en vez
    de usar la librería `mapbox_vector_tile`: esa librería devuelve los string_value con
    caracteres acentuados corruptos (mojibake) pese a que los bytes crudos del .pbf están en
    UTF-8 válido (verificado manualmente: b'Urucar\\xc3\\xa1' -> "Urucará"). Aquí se replica
    el esquema Tile/Layer/Feature/Value de la especificación Mapbox Vector Tile a mano.

    Devuelve (texto_extraido, es_pbf). Si falla la decodificación de un .pbf real, la
    excepción se relanza tal cual (ver manejo en `procesar_archivo`): no se asume
    silenciosamente que el archivo no tiene texto.
    """
    extension = ruta.suffix.lower()
    if extension != ".pbf":
        # .avif u otro formato no textual conocido en este corpus: no se intenta OCR/parseo,
        # se documenta explícitamente en vez de fallar.
        return "", False

    with open(ruta, "rb") as fh:
        datos_pbf = fh.read()

    mensaje, _ = blackboxprotobuf.decode_message(datos_pbf)
    capas = mensaje.get("3", [])
    if not isinstance(capas, list):
        capas = [capas]

    partes_texto = []
    for capa in capas:
        nombre_capa = _bytes_o_dict_a_str(capa.get("1", ""))

        claves = capa.get("3", [])
        if not isinstance(claves, list):
            claves = [claves]
        claves = [_bytes_o_dict_a_str(clave) for clave in claves]

        valores_crudos = capa.get("4", [])
        if not isinstance(valores_crudos, list):
            valores_crudos = [valores_crudos]
        valores = [_valor_mvt_a_texto(valor) for valor in valores_crudos]

        features = capa.get("2", [])
        if not isinstance(features, list):
            features = [features]

        for feature in features:
            tags_crudos = feature.get("2", b"")
            if not tags_crudos:
                continue
            indices = _decodificar_varints_empaquetados(tags_crudos)
            lineas = []
            for i in range(0, len(indices) - 1, 2):
                clave = claves[indices[i]]
                valor = valores[indices[i + 1]]
                lineas.append(f"{clave}: {valor}")
            if lineas:
                partes_texto.append(f"[{nombre_capa}] " + " | ".join(lineas))

    return "\n".join(partes_texto), True


# ---------------------------------------------------------------------------
# Orquestación por archivo y por fenómeno
# ---------------------------------------------------------------------------

DESPACHO_POR_TIPO = {
    "PDF": lambda ruta: extraer_pdf(ruta),
    "CSV": lambda ruta: extraer_csv(ruta),
    "Excel": lambda ruta: extraer_excel(ruta),
    "Texto": lambda ruta: extraer_texto_plano(ruta),
    "Imagen": lambda ruta: extraer_imagen(ruta),
}


def procesar_archivo(fila: FilaInventario, ruta: Path) -> DocumentoLimpio | DocumentoPendiente:
    """Despacha la extracción según el tipo declarado en el inventario y arma el registro final."""
    doc_id = f"{fila.fenomeno}-{fila.observatorio}-{fila.nombre_estandarizado}"

    try:
        if fila.tipo == "JSON":
            texto, metadata = extraer_json(ruta)
            if metadata.get("es_bulk_dump"):
                return DocumentoPendiente(
                    doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
                    fenomeno=fila.fenomeno, motivo="bulk_dump_json",
                    detalle=f"raíz es una lista con {metadata.get('num_registros')} registros",
                )
        elif fila.tipo == "Otro":
            texto, es_pbf = extraer_otro(ruta)
            if not es_pbf:
                return DocumentoPendiente(
                    doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
                    fenomeno=fila.fenomeno, motivo="tipo_otro_no_textual",
                    detalle=f"extensión {ruta.suffix} sin extractor de texto definido",
                )
        else:
            extractor = DESPACHO_POR_TIPO.get(fila.tipo)
            if extractor is None:
                return DocumentoPendiente(
                    doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
                    fenomeno=fila.fenomeno, motivo="tipo_sin_extractor",
                    detalle=f"tipo declarado en inventario: {fila.tipo}",
                )
            texto = extractor(ruta)
    except OCRNoDisponible as error:
        # Motivo distinguible del error genérico: se puede reprocesar solo estos, más
        # adelante, apenas se instalen Tesseract/Poppler, sin tocar el resto del pipeline.
        return DocumentoPendiente(
            doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
            fenomeno=fila.fenomeno, motivo="ocr_no_disponible",
            detalle=str(error),
        )
    except Exception as error:
        # Nunca se asume silenciosamente "sin texto": el error real queda registrado.
        return DocumentoPendiente(
            doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
            fenomeno=fila.fenomeno, motivo="error_extraccion",
            detalle=f"{type(error).__name__}: {error}",
        )

    if not texto.strip():
        return DocumentoPendiente(
            doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
            fenomeno=fila.fenomeno, motivo="texto_vacio_o_error",
            detalle="extracción completada sin errores pero el texto resultante está vacío",
        )

    documento = DocumentoLimpio(
        doc_id=doc_id,
        fuente=fila.observatorio,
        formato=fila.tipo,
        fenomeno=fila.fenomeno,
        texto_limpio=texto,
        num_caracteres=len(texto),
        idioma_detectado=detectar_idioma(texto),
    )

    if fila.tipo == "Otro" and ruta.suffix.lower() == ".pbf":
        documento.hash_contenido = calcular_hash(texto)
        # posible_duplicado se resuelve en una segunda pasada (ver `marcar_duplicados_pbf`)
        documento.posible_duplicado = False

    return documento


def clasificar_fila(fenomeno: str, fila: FilaInventario) -> DocumentoLimpio | DocumentoPendiente:
    """Resuelve y procesa una única fila del inventario (catálogo / no descargado / extracción)."""
    doc_id = f"{fila.fenomeno}-{fila.observatorio}-{fila.nombre_estandarizado}"

    if es_catalogo(fila.nombre_estandarizado):
        return DocumentoPendiente(
            doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
            fenomeno=fenomeno, motivo="catalogo_excluido",
            detalle="nombre coincide con patrón de catálogo/registro de metadata",
        )

    ruta = resolver_ruta_fisica(fila)
    if ruta is None:
        return DocumentoPendiente(
            doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
            fenomeno=fenomeno, motivo="no_descargado_localmente",
            detalle=f"carpeta esperada: {fila.carpeta}",
        )

    return procesar_archivo(fila, ruta)


# ---------------------------------------------------------------------------
# Checkpoints: hacen la extracción resiliente a crashes/cuelgues a mitad de corrida
# ---------------------------------------------------------------------------

class YaHayUnaCorridaEnCurso(Exception):
    """Se lanza si ya existe un lock activo para este fenómeno: dos procesos escribiendo
    a la vez sobre el mismo JSONL/checkpoint corrompen los datos (visto en producción:
    duplicados y PermissionError de Windows por archivos bloqueados entre sí)."""


def ruta_lock(fenomeno: str) -> Path:
    return CARPETA_SALIDA / f"{fenomeno.lower()}.lock"


class LockFenomeno:
    """
    Impide que dos corridas de `extraccion.py` procesen el mismo fenómeno al mismo tiempo.
    Usa creación exclusiva de archivo (O_CREAT|O_EXCL): si el lock ya existe, falla rápido
    en vez de dejar que dos procesos escriban al mismo JSONL/checkpoint a la vez.

    Si un proceso anterior murió sin limpiar el lock (crash duro, kill -9, corte de luz),
    el archivo queda huérfano: hay que confirmar que ya no hay ningún proceso corriendo y
    borrarlo a mano antes de reintentar. No se hace liveness-check automático del PID
    porque en Windows no es confiable sin dependencias externas.
    """

    def __init__(self, fenomeno: str):
        self.ruta = ruta_lock(fenomeno)

    def __enter__(self):
        CARPETA_SALIDA.mkdir(parents=True, exist_ok=True)
        try:
            fd = os.open(self.ruta, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            raise YaHayUnaCorridaEnCurso(
                f"Ya existe {self.ruta}: o hay otra corrida de este fenómeno en curso, o "
                f"quedó huérfano de un crash anterior. Verifica con el Administrador de "
                f"tareas que no haya ningún python.exe corriendo extraccion.py y, si "
                f"efectivamente no lo hay, borra el archivo a mano antes de reintentar."
            )
        with os.fdopen(fd, "w") as fh:
            fh.write(f"pid={os.getpid()}\n")
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.ruta.unlink(missing_ok=True)
        return False


def ruta_checkpoint(fenomeno: str) -> Path:
    return CARPETA_SALIDA / f"{fenomeno.lower()}_procesados.txt"


def ruta_documentos(fenomeno: str) -> Path:
    return CARPETA_SALIDA / f"{fenomeno.lower()}_documentos.jsonl"


def ruta_pendientes(fenomeno: str) -> Path:
    return CARPETA_SALIDA / f"{fenomeno.lower()}_pendientes.jsonl"


def cargar_checkpoint(fenomeno: str) -> set[str]:
    """Lee los doc_id ya completados en una corrida anterior, si el checkpoint existe."""
    ruta = ruta_checkpoint(fenomeno)
    if not ruta.exists():
        return set()
    with open(ruta, "r", encoding="utf-8") as fh:
        return {linea.strip() for linea in fh if linea.strip()}


def _corregir_duplicados_pbf(ruta_docs: Path, conteo_hash: dict[str, int]) -> None:
    """
    Segunda pasada ligera sobre el JSONL ya escrito: marca posible_duplicado=True en los
    registros cuyo hash_contenido se repite (varios .pbf con el mismo contenido). Se hace
    aquí, después de escribir todo en streaming, para no tener que mantener el texto
    completo de cada documento en memoria durante la extracción.
    """
    if not ruta_docs.exists():
        return

    fd_tmp, ruta_tmp = tempfile.mkstemp(dir=ruta_docs.parent, suffix=".tmp")
    with os.fdopen(fd_tmp, "w", encoding="utf-8") as f_tmp, open(ruta_docs, "r", encoding="utf-8") as f_docs:
        for linea in f_docs:
            registro = json.loads(linea)
            hash_contenido = registro.get("hash_contenido")
            if hash_contenido and conteo_hash.get(hash_contenido, 0) > 1:
                registro["posible_duplicado"] = True
            f_tmp.write(json.dumps(registro, ensure_ascii=False) + "\n")

    _reemplazar_con_reintento(ruta_tmp, ruta_docs)


def _reemplazar_con_reintento(origen: Path, destino: Path, intentos: int = 6) -> None:
    """
    os.replace puede fallar en Windows con PermissionError/WinError 5 si el archivo
    destino tiene un handle abierto momentáneamente (antivirus, indexador, u otro proceso,
    p.ej. OneDrive si el repo llegara a vivir dentro de una carpeta sincronizada).
    Reintenta con backoff antes de darse por vencido. Con 6 intentos y 1-2s por intento el
    tope es ~9s de espera acumulada, suficiente para que un lock momentáneo se libere (el
    crash observado en la prueba de 100 fue con el os.replace sin ningún reintento).
    """
    for intento in range(1, intentos + 1):
        try:
            os.replace(origen, destino)
            return
        except PermissionError:
            if intento == intentos:
                raise
            time.sleep(min(1.0 + 0.2 * intento, 2.0))


def _purgar_pendientes_reintentables(ruta_pend: Path, doc_ids_objetivo: set[str]) -> None:
    """
    Antes de reintentar, quita del jsonl de pendientes las líneas de doc_ids que están a
    punto de reprocesarse por un motivo transitorio (p.ej. ocr_no_disponible), para que no
    quede una entrada vieja duplicada u obsoleta si ahora sí se completan.
    """
    if not ruta_pend.exists():
        return

    with open(ruta_pend, "r", encoding="utf-8") as fh:
        lineas = fh.readlines()

    lineas_conservadas = []
    huno_cambio = False
    for linea in lineas:
        registro = json.loads(linea)
        if registro["doc_id"] in doc_ids_objetivo and registro.get("motivo") in MOTIVOS_REINTENTABLES:
            huno_cambio = True
            continue
        lineas_conservadas.append(linea)

    if huno_cambio:
        with open(ruta_pend, "w", encoding="utf-8") as fh:
            fh.writelines(lineas_conservadas)


def procesar_fenomeno_streaming(
    fenomeno: str, filas: list[FilaInventario], limite: int | None = None
) -> tuple[int, int]:
    """
    Procesa las filas de un fenómeno documento por documento, escribiendo cada resultado
    en el JSONL correspondiente de inmediato (append + flush) y registrando el doc_id en
    el checkpoint apenas termina. Si el proceso se cuelga o se cierra, al reiniciar se
    salta todo lo que ya está en el checkpoint y continúa donde quedó.

    Los pendientes por un motivo transitorio (MOTIVOS_REINTENTABLES, p.ej. falta de
    Tesseract/Poppler) NO se marcan como completados: se reintentan automáticamente en la
    siguiente corrida en vez de quedar perdidos para siempre.

    Devuelve (documentos_nuevos, pendientes_nuevos) escritos en ESTA corrida.

    Protegido con LockFenomeno: si ya hay otra corrida de este mismo fenómeno en curso
    (p.ej. un proceso en segundo plano de un intento anterior que no murió), lanza
    YaHayUnaCorridaEnCurso en vez de dejar que dos procesos escriban a la vez sobre el
    mismo JSONL/checkpoint (eso ya corrompió datos una vez: duplicados + PermissionError).
    """
    with LockFenomeno(fenomeno):
        return _procesar_fenomeno_streaming_interno(fenomeno, filas, limite)


def _procesar_fenomeno_streaming_interno(
    fenomeno: str, filas: list[FilaInventario], limite: int | None
) -> tuple[int, int]:
    ya_procesados = cargar_checkpoint(fenomeno)
    filas_objetivo = filas[:limite] if limite is not None else filas
    total = len(filas_objetivo)

    CARPETA_SALIDA.mkdir(parents=True, exist_ok=True)
    ruta_docs = ruta_documentos(fenomeno)
    ruta_pend = ruta_pendientes(fenomeno)
    ruta_chk = ruta_checkpoint(fenomeno)

    doc_ids_objetivo = {
        f"{fila.fenomeno}-{fila.observatorio}-{fila.nombre_estandarizado}" for fila in filas_objetivo
    }
    _purgar_pendientes_reintentables(ruta_pend, doc_ids_objetivo)

    conteo_hash: dict[str, int] = {}
    docs_nuevos = 0
    pend_nuevos = 0

    with open(ruta_docs, "a", encoding="utf-8") as f_docs, \
         open(ruta_pend, "a", encoding="utf-8") as f_pend, \
         open(ruta_chk, "a", encoding="utf-8") as f_chk:

        for i, fila in enumerate(filas_objetivo, start=1):
            doc_id = f"{fila.fenomeno}-{fila.observatorio}-{fila.nombre_estandarizado}"

            if doc_id in ya_procesados:
                continue

            print(f"{i}/{total}  {doc_id}  ({fila.nombre_estandarizado})", flush=True)

            try:
                verificar_memoria_o_pausar()
                resultado = clasificar_fila(fenomeno, fila)
            except MemoriaInsuficiente as error:
                resultado = DocumentoPendiente(
                    doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
                    fenomeno=fenomeno, motivo="memoria_insuficiente",
                    detalle=str(error),
                )
            except Exception as error:
                # Ningún documento individual puede tumbar el proceso completo.
                resultado = DocumentoPendiente(
                    doc_id=doc_id, fuente=fila.observatorio, formato=fila.tipo,
                    fenomeno=fenomeno, motivo="error_inesperado",
                    detalle=f"{type(error).__name__}: {error}",
                )

            es_reintentable = False
            if isinstance(resultado, DocumentoLimpio):
                if resultado.hash_contenido:
                    conteo_hash[resultado.hash_contenido] = conteo_hash.get(resultado.hash_contenido, 0) + 1
                f_docs.write(json.dumps(resultado.to_dict(), ensure_ascii=False) + "\n")
                f_docs.flush()
                docs_nuevos += 1
            else:
                es_reintentable = resultado.motivo in MOTIVOS_REINTENTABLES
                f_pend.write(json.dumps(resultado.to_dict(), ensure_ascii=False) + "\n")
                f_pend.flush()
                pend_nuevos += 1

            if not es_reintentable:
                f_chk.write(doc_id + "\n")
                f_chk.flush()
                ya_procesados.add(doc_id)

            if i % INTERVALO_GC == 0:
                gc.collect()

    if conteo_hash:
        _corregir_duplicados_pbf(ruta_docs, conteo_hash)

    return docs_nuevos, pend_nuevos


# ---------------------------------------------------------------------------
# Validación contra el resumen del Excel
# ---------------------------------------------------------------------------

def leer_resumen_por_fenomeno(ruta_indice: Path = RUTA_INDICE) -> dict[str, dict[str, int]]:
    """Lee la hoja 'Resumen por Fenomeno' (Total datos, PDFs, JSONs) como referencia de validación."""
    wb = openpyxl.load_workbook(ruta_indice, data_only=True)
    ws = wb["Resumen por Fenomeno"]
    filas = list(ws.iter_rows(values_only=True))
    encabezado = filas[0]
    idx = {nombre: i for i, nombre in enumerate(encabezado)}

    resumen = {}
    for fila in filas[1:]:
        fenomeno_completo = fila[idx["Fenómeno"]]
        if not fenomeno_completo or not str(fenomeno_completo).startswith("F"):
            continue
        codigo = str(fenomeno_completo).split()[0]  # "F1 – ..." -> "F1"
        resumen[codigo] = {
            "total_datos": fila[idx["Total datos"]],
            "pdfs": fila[idx["PDFs"]],
            "jsons": fila[idx["JSONs"]],
        }
    return resumen


def _contar_lineas(ruta: Path) -> int:
    if not ruta.exists():
        return 0
    with open(ruta, "r", encoding="utf-8") as fh:
        return sum(1 for _ in fh)


def validar_conteos(fenomenos: Iterable[str], inventario: list[FilaInventario]) -> list[str]:
    """
    Compara, para cada fenómeno, los conteos del Inventario (esperado) contra lo realmente
    escrito en disco (documentos limpios + pendientes) y contra la hoja "Resumen por
    Fenomeno". Lee los JSONL de salida en vez de listas en memoria, porque en una corrida
    parcial (checkpoint) lo único confiable es lo que ya quedó escrito en el disco.
    """
    resumen_excel = leer_resumen_por_fenomeno()
    reporte = []

    for fenomeno in fenomenos:
        total_inventario = sum(1 for f in inventario if f.fenomeno == fenomeno)
        num_documentos = _contar_lineas(ruta_documentos(fenomeno))
        num_pendientes = _contar_lineas(ruta_pendientes(fenomeno))
        total_procesado = num_documentos + num_pendientes

        linea = (
            f"{fenomeno}: inventario={total_inventario} | "
            f"procesados+pendientes={total_procesado} | "
            f"documentos_limpios={num_documentos} | pendientes={num_pendientes}"
        )
        if fenomeno in resumen_excel:
            linea += f" | resumen_excel_total_datos={resumen_excel[fenomeno]['total_datos']}"
        if total_inventario != total_procesado:
            linea += "  <-- DESCUADRE (puede ser normal si es una corrida parcial con --limite)"
        reporte.append(linea)

    return reporte


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def ejecutar(
    fenomenos_a_procesar: Iterable[str] = FENOMENOS,
    limite: int | None = None,
    muestra_mixta: bool = False,
) -> None:
    """
    Orquesta la extracción: lee el inventario y procesa cada fenómeno solicitado en modo
    streaming (append + checkpoint por documento).

    - `limite`: solo procesa las primeras N filas del inventario del fenómeno (o, si
      `muestra_mixta=True`, N filas repartidas entre tipos). None procesa todo lo pendiente.
    - `muestra_mixta`: usa `seleccionar_muestra_mixta` en vez de tomar las primeras N filas
      tal cual — pensado para las pruebas iniciales de 20/100 documentos.
    """
    inventario = leer_inventario()

    for fenomeno in fenomenos_a_procesar:
        filas_fenomeno = [f for f in inventario if f.fenomeno == fenomeno]
        if limite is not None and muestra_mixta:
            filas_fenomeno = seleccionar_muestra_mixta(filas_fenomeno, limite)
            docs_nuevos, pend_nuevos = procesar_fenomeno_streaming(fenomeno, filas_fenomeno, limite=None)
        else:
            docs_nuevos, pend_nuevos = procesar_fenomeno_streaming(fenomeno, filas_fenomeno, limite=limite)
        print(f"-- {fenomeno}: {docs_nuevos} documentos nuevos, {pend_nuevos} pendientes nuevos en esta corrida --")

    print()
    for linea in validar_conteos(fenomenos_a_procesar, inventario):
        print(linea)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extracción del corpus CODEFEST AD ASTRA 2026")
    parser.add_argument(
        "--fenomenos", nargs="+", default=list(FENOMENOS), choices=list(FENOMENOS),
        help="Fenómenos a procesar (default: todos)",
    )
    parser.add_argument(
        "--limite", type=int, default=None,
        help="Procesar solo N filas del inventario por fenómeno (pruebas)",
    )
    parser.add_argument(
        "--muestra-mixta", action="store_true",
        help="Con --limite, reparte las N filas entre los tipos presentes (PDF/JSON/CSV/Otro) "
             "en vez de tomar las primeras N tal cual",
    )
    args = parser.parse_args()
    ejecutar(fenomenos_a_procesar=args.fenomenos, limite=args.limite, muestra_mixta=args.muestra_mixta)
